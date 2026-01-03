import logging
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class KapitalBankAPI:
    """Клиент для работы с API Kapital Bank и сохранения отчёта в Excel (Accounts, Statements, POS Operations)"""

    def __init__(self, excel_path: Path):

        self.excel_path = excel_path

        self.base_url = "https://my.birbank.business/api/b2b"
        self.clientId = ""
        self.refreshToken = ""
        self.token = ""
        self.accounts = []
        self.statements_dataset = []
        self.cards_statements = []
        self.cards = []

        self.session = requests.Session()

    def _authenticate(self, username: str, password: str):
        try:
            response = self.session.post(
                f"{self.base_url}/login",
                json={"username": username, "password": password},
                headers={"User-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                         " (KHTML, like Gecko) Chrome/102.0.5005.49 Safari/537.36"},
                timeout=30,
            )
            response.raise_for_status()
            logging.info("Authentication request sent successfully.")
        except requests.RequestException as e:
            logging.error(f"Authentication request failed: {e}")
            return False

        try:
            user_data = response.json()
            logging.info("Authentication response parsed successfully.")
        except ValueError as e:
            logging.error(f"Failed to parse authentication response JSON: {e}")
            return False

        response_data = user_data.get("responseData") or {}
        user_info = (response_data.get("userInfo") or {})
        chat_data = (user_info.get("chatData") or {})

        client_id = chat_data.get("clientId")
        refresh_token = response_data.get("jwtrefreshtoken")
        token = response_data.get("jwttoken")

        if client_id and refresh_token and token:
            self.clientId = client_id
            self.refreshToken = refresh_token
            self.token = token

            self.session.headers.update({"Authorization": f"Bearer {self.token}"})
            self.session.headers["User-agent"] = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                                                  " (KHTML, like Gecko) Chrome/102.0.5005.49 Safari/537.36")

            logging.info("User authenticated successfully!")
            return True
        else:
            logging.warning("Authentication tokens missing in response.")
            return False

    def _get_accounts(self):
        try:
            logging.info("Getting accounts")
            response = self.session.get(url=f"{self.base_url}/accounts")
            response.raise_for_status()
            data = response.json()
            self.accounts = data.get("responseData", {}).get("accountsList", [])
            logging.info(f"Accounts retrieved successfully. Number of accounts: {len(self.accounts)}")
            logging.info(self.accounts)

        except requests.RequestException as e:
            logging.error(f"Failed to get accounts: {e}")
            self.accounts = []

    def _get_statements_for_accounts(self, date_from: str, date_to: str):
        self._get_accounts()

        logging.info("Processing each account's statements:")

        for account in self.accounts:
            logging.info(f"Processing account: {account.get('custAcNo')}")

            try:

                response = self.session.get(f"{self.base_url}/v2/statement/account?fromDate={date_from}"
                                            f"&toDate={date_to}&accountNumber={account.get('custAcNo')}")

                logging.info(f"Getting statements for account {account.get('custAcNo')}")
                response.raise_for_status()
                data = response.json()
                logging.info(f"Statements retrieved successfully for account {account.get('custAcNo')}")

                self.statements_dataset.append(data)

            except requests.RequestException as e:
                logging.error(f"Failed to get statements for account {account.get('custAcNo')}: {e}")


    def _get_cards_data(self) -> list:

        for account in self.accounts:
            logging.info(f"Getting cards data for account: {account.get('custAcNo')}")

            try:
                response = self.session.get(f"{self.base_url}/cards")
                response.raise_for_status()

                logging.info(f"Cards data retrieved successfully for account {account.get('custAcNo')}")

                cards_data = response.json().get("responseData", {}).get("cards", [])

                self.cards.extend(cards_data)

            except requests.RequestException as e:
                logging.error(f"Failed to get cards data for account {account.get('custAcNo')}: {e}")

        return self.cards

    def _calculate_90_days_period(self, start_date, end_date) -> list[dict[str, str]]:

        start_obj = datetime.strptime(start_date, "%Y-%m-%d")
        end_obj = datetime.strptime(end_date, "%Y-%m-%d")

        month_selected = ((end_obj - start_obj).days // 30)

        prev_date = ""
        next_date = ""

        date_containing = []

        if month_selected >= 3:
            while month_selected > 0:

                if prev_date == "":
                    prev_date = start_date

                if next_date != "":
                    prev_date = next_date

                if (datetime.strptime(prev_date, "%Y-%m-%d") + timedelta(days=89)) > datetime.now():
                    break

                next_date = (datetime.strptime(prev_date, "%Y-%m-%d") + timedelta(days=89)).strftime(
                    "%Y-%m-%d")

                date_containing.append({"start": prev_date, "end": next_date})

                month_selected -= 3
        else:
            date_containing.append({"start": start_date, "end": end_date})

        return date_containing



    def _get_cards_statements(self, from_date: str, to_date: str):

        cards_data = self._get_cards_data()
        logging.info(f"Cards data retrieved successfully. Number of cards: {len(cards_data)}")
        logging.info(cards_data)

        date_parts = from_date.split("-")
        date_parts_2 = to_date.split("-")

        new_from_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
        new_to_date = f"{date_parts_2[2]}-{date_parts_2[1]}-{date_parts_2[0]}"

        date_objects = self._calculate_90_days_period(new_from_date, new_to_date)

        if not cards_data:
            logging.warning("No cards found to get statements for.")
            return

        if not date_objects:
            logging.warning("No periods found to get statements for.")
            return


        for card in cards_data:
            logging.info(f"Getting cards statements for card account: {card.get('accountNumber')}")

            for period in date_objects:

                logging.info(f"Getting cards statements for period: {period.get('start')} - {period.get('end')}")

                try:
                    response = self.session.get(f"{self.base_url}/v2/statement/card?fromDate={period.get('start')}"
                                                f"&toDate={period.get("end")}&accountNumber={card.get('accountNumber')}")

                    response.raise_for_status()
                    data = response.json()

                    dataset = data.get("responseData", {}).get("operation", [])

                    if not dataset:
                        logging.warning(f"No statements found for account {card.get('accountNumber')}")
                        continue
                    else:
                        logging.info(f"Cards statements retrieved successfully for account {card.get('accountNumber')}")

                    self.cards_statements.extend(dataset)

                except requests.RequestException as e:
                    logging.error(f"Failed to get cards statements for card account {card.get('accountNumber')}: {e}")

        logging.info(f"Cards statements retrieved successfully. Number of statements: {len(self.cards_statements)}")
        logging.info(self.cards_statements)

    def _prepare_excel(self):
        accounts_table = []

        if self.accounts:
            for account in self.accounts:
                accounts_table.append({
                    "Branch Code": account.get("branchCode", ""),
                    "Customer Account No": account.get("custAcNo", ""),
                    "IBAN Account No": account.get("ibanAcNo", ""),
                    "Currency": account.get("ccy", ""),
                    "Status": account.get("status", ""),
                    "Planned Amount": account.get("plannedAmt", ""),
                    "Current Amount": account.get("currAmt", ""),
                    "Hold": account.get("hold", ""),
                })

        wb = Workbook()
        ws_acc = wb.active
        ws_acc.title = "Accounts"

        if accounts_table:
            df_accounts = pd.DataFrame(accounts_table)
            for r in dataframe_to_rows(df_accounts, index=False, header=True):
                ws_acc.append(r)

            for cell in ws_acc[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            ws_acc.auto_filter.ref = ws_acc.dimensions
            for column_cells in ws_acc.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value)
                    except Exception:
                        cell_value = ""
                    max_length = max(max_length, len(cell_value))
                adjusted_width = (max_length + 2)
                ws_acc.column_dimensions[column].width = adjusted_width
            logging.info("Excel sheet prepared with account data.")

        else:
            ws_acc["A1"] = "No accounts found"
            logging.warning("No accounts found to write to Excel.")

        date_suffix = datetime.now().strftime("%Y-%m-%d_%H-%M")
        final_filename = f"{date_suffix}_kapital_report.xlsx"

        if self.excel_path:
            final_filename = str(self.excel_path.joinpath(final_filename))
            logging.log(msg="Final path: " + final_filename, level=logging.INFO)

        if self.statements_dataset:
            statements_sheet = wb.create_sheet("Accounts_Statements")
            current_row = 1

            for dataset in self.statements_dataset:
                try:

                    account_info = dataset.get("responseData", {}).get("operations", {}).get("accountInfo", {})
                    statements = dataset.get("responseData", {}).get("operations", {}).get("statementList", [])

                    if not account_info:
                        logging.warning("No account info found in dataset")
                        continue

                    # ===== ДОБАВЛЯЕМ ИНФОРМАЦИЮ АККАУНТА =====
                    statements_sheet[f"A{current_row}"] = "=== ACCOUNT INFO ==="
                    statements_sheet[f"A{current_row}"].font = Font(bold=True, size=12, color="000000")
                    statements_sheet[f"A{current_row}"].fill = PatternFill(start_color="366092", end_color="366092",
                                                                           fill_type="solid")
                    current_row += 1

                    # Заголовки для информации аккаунта
                    account_headers = list(account_info.keys())
                    for idx, header in enumerate(account_headers, start=1):
                        col_letter = chr(64 + idx)  # A, B, C, ...
                        statements_sheet[f"{col_letter}{current_row}"] = header
                        statements_sheet[f"{col_letter}{current_row}"].font = Font(bold=True)
                        statements_sheet[f"{col_letter}{current_row}"].fill = PatternFill(start_color="BDD7EE",
                                                                                          end_color="BDD7EE",
                                                                                          fill_type="solid")
                    current_row += 1

                    # Значения информации аккаунта
                    for idx, header in enumerate(account_headers, start=1):
                        col_letter = chr(64 + idx)
                        statements_sheet[f"{col_letter}{current_row}"] = account_info.get(header, "")
                    current_row += 2  # Пустая строка между аккаунтом и операциями

                    # ===== ДОБАВЛЯЕМ ОПЕРАЦИИ =====
                    if statements:
                        statements_sheet[f"A{current_row}"] = "=== STATEMENT LIST ==="
                        statements_sheet[f"A{current_row}"].font = Font(bold=True, size=12, color="000000")
                        statements_sheet[f"A{current_row}"].fill = PatternFill(start_color="366092", end_color="366092",
                                                                               fill_type="solid")
                        current_row += 1

                        df_statements = pd.DataFrame(statements)

                        # Заголовки операций
                        statement_headers = list(df_statements.columns)
                        for idx, header in enumerate(statement_headers, start=1):
                            col_letter = chr(64 + idx)
                            statements_sheet[f"{col_letter}{current_row}"] = header
                            statements_sheet[f"{col_letter}{current_row}"].font = Font(bold=True)
                            statements_sheet[f"{col_letter}{current_row}"].fill = PatternFill(start_color="BDD7EE",
                                                                                              end_color="BDD7EE",
                                                                                              fill_type="solid")
                        current_row += 1

                        # Значения операций
                        for _, row in df_statements.iterrows():
                            for idx, header in enumerate(statement_headers, start=1):
                                col_letter = chr(64 + idx)
                                statements_sheet[f"{col_letter}{current_row}"] = row[header]
                            current_row += 1

                        current_row += 2  # Пустые строки между аккаунтами

                    logging.info(f"Added account info and {len(statements)} statements to Excel")

                except Exception as e:
                    logging.error(f"Error processing dataset: {e}")
                    continue

            # Авторазмер колонок в листе Statements
            for column_cells in statements_sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value)
                    except Exception:
                        cell_value = ""
                    max_length = max(max_length, len(cell_value))
                adjusted_width = min(max_length + 2, 50)
                statements_sheet.column_dimensions[column].width = adjusted_width

        if self.cards:
            cards_sheet = wb.create_sheet("Cards")
            counter = 1

            # Заголовки (первая строка)
            for idx, card_header in enumerate(self.cards[0].keys(), start=1):
                col_letter = chr(64 + idx)  # A, B, C, ...
                cards_sheet[f"{col_letter}1"] = card_header
                cards_sheet[f"{col_letter}1"].font = Font(bold=True)
                cards_sheet[f"{col_letter}1"].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE",
                                                                 fill_type="solid")

            counter += 1  # увеличить строку после заголовков один раз

            # Данные карт (начиная со второй строки)
            for card in self.cards:
                for val_id, card_val in enumerate(card.values(), start=1):
                    letter = chr(64 + val_id)
                    cards_sheet[f"{letter}{counter}"] = card_val
                    cards_sheet[f"{letter}{counter}"].font = Font(size=12, color="000000")

                counter += 1

            # Авторазмер колонок
            for column_cells in cards_sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value)
                    except Exception:
                        cell_value = ""
                    max_length = max(max_length, len(cell_value))
                adjusted_width = min(max_length + 2, 50)
                cards_sheet.column_dimensions[column].width = adjusted_width

            logging.info(f"Cards sheet created with {len(self.cards)} cards")

        if self.cards_statements:
            cards_statements_sheet = wb.create_sheet("Cards_Statements")

            # Заголовки из ключей первого словаря
            headers = list(self.cards_statements[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                col_letter = chr(64 + col_idx)
                cell = cards_statements_sheet[f"{col_letter}1"]
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Данные операций
            for row_idx, operation in enumerate(self.cards_statements, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    col_letter = chr(64 + col_idx)
                    val = operation.get(header, "")
                    cell = cards_statements_sheet[f"{col_letter}{row_idx}"]
                    cell.value = val
                    cell.font = Font(size=11, color="000000")

            # Авторазмер колонок
            for column_cells in cards_statements_sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value)
                    except Exception:
                        cell_value = ""
                    max_length = max(max_length, len(cell_value))
                adjusted_width = min(max_length + 2, 50)
                cards_statements_sheet.column_dimensions[column].width = adjusted_width

            logging.info(f"Card Statements sheet created with {len(self.cards_statements)} operations")

        date_suffix = datetime.now().strftime("%Y-%m-%d_%H-%M")
        final_filename = f"{date_suffix}_kapital_report.xlsx"

        if self.excel_path:
            final_filename = str(self.excel_path.joinpath(final_filename))
            logging.info("Final path: " + final_filename)

        wb.save(final_filename)
        logging.info(f"Excel file saved as {final_filename}")

        return True

    def process_data(self, date_from: str, date_to: str, username: str, password: str):
        #аутентифицировать перед запросами
        self._authenticate(username, password)

        logging.info(f"Process data called with date_from={date_from} and date_to={date_to}")

        if not date_from or not date_to:
            logging.error("Date range is not valid. Please check your input.")
            return False

        self._get_statements_for_accounts(date_from, date_to)
        self._get_cards_statements(date_from, date_to)
        return self._prepare_excel()
