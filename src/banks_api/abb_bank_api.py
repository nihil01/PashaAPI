import logging
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Optional

class AbbBankAPI:
    """
    Client for ABB Bank API interactions.
    """

    def __init__(self, excel_path: Path) -> None:
        self.base_url = "https://api-prod-c2b.abb-bank.az"
        self.session = requests.Session()
        self.token: Optional[str] = None
        self.excel_path = excel_path

        self.accounts = []
        self.statements_dataset = []

    def authenticate(self, username: str, password: str) -> bool:
        url = f"{self.base_url}/payments/auth/token"
        
        headers = {
            "Content-Type": "application/json",
            "Charset": "UTF-8"
        }
        
        payload = {
            "username": username,
            "password": password
        }

        try:
            logging.info(f"Attempting to authenticate with ABB Bank at {url}")
            response = self.session.post(url, json=payload, headers=headers, timeout=30)
            response.raise_for_status()

            data = response.json()
            logging.info(data)
            
            # Attempt to retrieve token from common fields since response schema isn't fully specified
            self.token = data.get("token") or data.get("access_token") or data.get("accessToken")

            if self.token:
                self._update_session_headers()
                logging.info("ABB Bank authentication successful.")
                return True
            else:
                logging.error(f"Authentication failed. Token not found in response: {data}")
                return False

        except requests.RequestException as e:
            logging.error(f"ABB Bank authentication error: {e}")
            return False

    def _update_session_headers(self):
        """Updates the session headers with the Bearer token."""
        if self.token:
            self.session.headers["Authorization"] = f"Bearer {self.token}"
            self.session.headers["Content-Type"] = "application/json"
            # Adding User-Agent similar to other clients in the project
            self.session.headers["User-agent"] = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                                                  "Chrome/102.0.5005.49 Safari/537.36")


    def get_accounts(self) -> list[dict[str, str]]:
        """
        Fetches the list of corporate accounts.

        Documentation:
            Method: GET
            URL: https://api-test-c2b.abb-bank.az/payments/corporate-account-info
        """
        url = f"{self.base_url}/payments/corporate-account-info"
        try:
            logging.info(f"Fetching accounts from {url}")
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            raw_accounts = response.json()

            if raw_accounts is not None:
                for account in raw_accounts:
                    self.accounts.append(account)


        except requests.RequestException as e:
            logging.error(f"Error fetching accounts: {e}")
            return []


    def get_account_statement(self, date_from: str, date_to: str,
                                  page: int = 1, page_size: int = 100) -> list[dict[str, str]]:
        """
        Fetches account statements.

        Documentation:
            Method: GET
            URL: https://api-test-c2b.abb-bank.az/payments/account-statement
            Query parameters: account, from-date, to-date, page-size, page, operation-type

            date_from, date_to: YYYY-MM-DD
            operation-type: A / D / C, default A
        """
        url = f"{self.base_url}/payments/account/statement"

        periods = self._calculate_30_days_period(date_from, date_to)

        for account in self.accounts:

            account_number = account.get("accountNo")

            for period in periods:

                date_from_p = period["start"]
                date_to_p = period["end"]

                params = {
                    "account": account_number,
                    "from-date": date_from_p,
                    "to-date": date_to_p,
                    "page-size": page_size,
                    "page": page
                }

                try:
                    logging.info(f"Fetching statements for {account_number} ({date_from_p} to {date_to_p})")
                    response = self.session.get(url, params=params, timeout=30)
                    data = response.json()
                    if data:
                        self.statements_dataset.append(data)

                except requests.RequestException as e:
                    logging.error(f"Error fetching statements: {e}")

        return self.statements_dataset

    def _calculate_30_days_period(self, start_date, end_date) -> list[dict[str, str]]:

        start_obj = datetime.strptime(start_date, "%Y-%m-%d")
        end_obj = datetime.strptime(end_date, "%Y-%m-%d")

        month_selected = ((end_obj - start_obj).days // 30)

        prev_date = ""
        next_date = ""

        date_containing = []

        if month_selected > 1:
            while month_selected > 0:

                if prev_date == "":
                    prev_date = start_date

                if next_date != "":
                    prev_date = next_date

                if (datetime.strptime(prev_date, "%Y-%m-%d") + timedelta(days=30)) > datetime.now():
                    break

                next_date = (datetime.strptime(prev_date, "%Y-%m-%d") + timedelta(days=30)).strftime("%Y-%m-%d")

                date_containing.append({"start": prev_date, "end": next_date})

                month_selected -= 1
        else:
            date_containing.append({"start": start_date, "end": end_date})

        return date_containing

    def _prepare_excel(self):
        wb = Workbook()

        standard_font = Font(name="Segoe UI", size=10)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        standard_alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)

        def apply_cell_style(cell, is_header=False, is_subheader=False):
            cell.font = header_font if is_header else standard_font
            cell.alignment = standard_alignment
            if is_header:
                cell.fill = header_fill
            elif is_subheader:
                cell.fill = subheader_fill

        # 1. Лист ACCOUNTS
        ws_accounts = wb.active
        ws_accounts.title = "Accounts"

        if not self.accounts:
            ws_accounts["A1"] = "No accounts found"
            logging.warning("No accounts found to write to Excel.")
        else:
            columns_accounts = [
                "name", "accountNo", "iban", "currency",
                "availableBalance", "todayOpeningBalance",
                "todayIncome", "todayOutcome"
            ]

            # Заголовки
            for idx, header in enumerate(columns_accounts, start=1):
                cell = ws_accounts.cell(row=1, column=idx, value=header.upper())
                apply_cell_style(cell, is_header=True)

            # Данные
            for row_idx, account in enumerate(self.accounts, start=2):
                for col_idx, header in enumerate(columns_accounts, start=1):
                    val = account.get(header)
                    # Обработка вложенного объекта branch, если нужно (в коде был accountBranch)
                    if header == "accountBranch" and isinstance(val, dict):
                        val = val.get("branchName")

                    cell = ws_accounts.cell(row=row_idx, column=col_idx, value=val)
                    apply_cell_style(cell)



        ws_statements = wb.create_sheet("Account_Statements")
        if not self.statements_dataset:
            ws_statements["A1"] = "No statements found"
            logging.warning("No statements found to write to Excel.")
            return False

        current_row = 1

        for dataset in self.statements_dataset:
            try:
                # ABB API response structure: {'accountInfo': {...}, 'transaction': {'transactions': [...], ...}}
                account_info = dataset.get("accountInfo", {})
                transaction_summary = dataset.get("transaction", {})
                transactions = transaction_summary.get("transactions", [])

                # 1. DYNAMIC ACCOUNT & SUMMARY INFO
                ws_statements.cell(row=current_row, column=1, value="=== ACCOUNT & SUMMARY INFO ===")
                ws_statements[f"A{current_row}"].font = Font(bold=True, size=12, color="FFFFFF")
                ws_statements[f"A{current_row}"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                current_row += 1

                # Flatten nested account info (like branch details) and transaction summaries
                combined_info = {}

                def flatten_dict(d, prefix=''):
                    for k, v in d.items():
                        if k == 'transactions': continue
                        if isinstance(v, dict):
                            flatten_dict(v, f"{prefix}{k}_")
                        else:
                            combined_info[f"{prefix}{k}"] = v

                flatten_dict(account_info)
                flatten_dict(transaction_summary)

                if combined_info:
                    headers = list(combined_info.keys())
                    for idx, header in enumerate(headers, start=1):
                        cell = ws_statements.cell(row=current_row, column=idx, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

                    # current_row += 1
                    for idx, header in enumerate(headers, start=1):
                        ws_statements.cell(row=current_row, column=idx, value=str(combined_info[header]))

                    current_row += 2

                # 2. DYNAMIC TRANSACTION LIST
                if transactions:
                    ws_statements.cell(row=current_row, column=1, value="=== TRANSACTION LIST ===")
                    ws_statements[f"A{current_row}"].font = Font(bold=True, size=12, color="FFFFFF")
                    ws_statements[f"A{current_row}"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    current_row += 1

                    df_trans = pd.DataFrame(transactions)
                    trans_headers = list(df_trans.columns)

                    # Write Headers
                    for idx, header in enumerate(trans_headers, start=1):
                        cell = ws_statements.cell(row=current_row, column=idx, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

                    current_row += 1

                    # Write Data
                    for _, row in df_trans.iterrows():
                        for idx, header in enumerate(trans_headers, start=1):
                            ws_statements.cell(row=current_row, column=idx, value=row[header])
                        current_row += 1

                    current_row += 2  # Gap between different account reports

            except Exception as e:
                logging.error(f"Error processing dataset for Excel: {e}")
                continue

            # Auto-size columns based on content
            for column_cells in ws_statements.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        if len(val) > max_length: max_length = len(val)
                    except: pass
                ws_statements.column_dimensions[column].width = min(max_length + 3, 60)

        date_suffix = datetime.now().strftime("%Y-%m-%d_%H-%M")
        final_filename = f"{date_suffix}_abb_report.xlsx"
        final_path = self.excel_path / final_filename if self.excel_path else Path(final_filename)

        wb.save(str(final_path))
        logging.info(f"Excel file saved as {final_path}")
        return True

    def process_data(self, date_from: str, date_to: str, username: str, password: str):
        # аутентифицировать перед запросами
        if not self.token:
            if not self.authenticate(username, password):
                logging.error("Authentication failed.")
                return False

        logging.info("Already authenticated.")
        logging.info(f"Process data called with date_from={date_from} and date_to={date_to}")

        if not date_from or not date_to:
            logging.error("Date range is not valid. Please check your input.")
            return False

        self.get_accounts()
        logging.info(f"Fetched {len(self.accounts)} accounts")

        if self.accounts:
            self.get_account_statement(date_from, date_to)
            return self._prepare_excel()

        #cleanup
        self.accounts = []
        self.statements_dataset = []


        return False
