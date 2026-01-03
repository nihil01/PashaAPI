import time

import requests
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path
import logging


def _normalize_value(v: Any) -> Any:
    """Replace None/empty string with N/A, keep numbers as-is."""
    if v is None:
        return "N/A"
    if isinstance(v, str) and v.strip() == "":
        return "N/A"
    return v


class PashaBankAPI:
    """Клиент для работы с API Pasha Bank и сохранения отчёта в Excel (Accounts, Statements, POS Operations)"""

    def __init__(self, excel_path: Path) -> None:

        self.excel_path = excel_path
        self.config_jwt = ""
        self.config_key = ""

        self.page_max_count = 0
        self.current_page = 0

        self.base_url = "https://openapi.pashabank.digital"
        self.accounts_list_path = "/api/v1/accounts"
        self.pos_operations = "/api/v1/accounts/{accountId}/statements/pos"
        self.stmt_path = "/api/v1/accounts/{accountId}/current/paginated"

        self.session = requests.Session()
        self._setup_session()


    def save_report(self, accounts_table: List[Dict[str, Any]],
                    statements_rows: List[Dict[str, Any]],
                    pos_rows: List[Dict[str, Any]],
                    filename="report.xlsx"):
        wb = Workbook()

        # Accounts sheet
        ws_acc = wb.active
        ws_acc.title = "Accounts"
        if accounts_table:
            df_accounts = pd.DataFrame(accounts_table)
            for r in dataframe_to_rows(df_accounts, index=False, header=True):
                ws_acc.append(r)
            for cell in ws_acc[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            ws_acc.auto_filter.ref = ws_acc.dimensions
        else:
            ws_acc["A1"] = "No accounts found"

        # Statements sheet
        ws_stmt = wb.create_sheet("Statements")
        if statements_rows:
            df_stmt = pd.DataFrame(statements_rows)
            preferred_order = [
                "accountNo",
                "operationDate", "transactionDate", "transactionNo", "transactionType",
                "transactionDescription",
                "amountInTransactionCurrency", "transactionCurrency",
                "amountInAccountCurrency", "amountInTransactionCurrencyAzn", "transactionFXRate",
                "openingBalance_op", "closingBalance_op", "openingBalance", "closingBalance",
                "availableOpeningBalance", "availableClosingBalance",
                "counterPartyName", "counterPartyId", "counterPartyTin",
                "cardNo", "sourceSystem", "message", "page_current", "page_total"
            ]
            cols = [c for c in preferred_order if c in df_stmt.columns] + [c for c in df_stmt.columns if
                                                                           c not in preferred_order]
            df_stmt = df_stmt[cols]
            for r in dataframe_to_rows(df_stmt, index=False, header=True):
                ws_stmt.append(r)
            for cell in ws_stmt[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            ws_stmt.auto_filter.ref = ws_stmt.dimensions
        else:
            ws_stmt["A1"] = "No statements found"

        # POS sheet (hybrid B1: summary row then operation rows)
        ws_pos = wb.create_sheet("POS Operations")
        if pos_rows:
            df_pos = pd.DataFrame(pos_rows)
            # prefer 'rowType' first
            cols = ["rowType"] + [c for c in df_pos.columns if c != "rowType"]
            df_pos = df_pos[cols]
            for r in dataframe_to_rows(df_pos, index=False, header=True):
                ws_pos.append(r)
            # style header
            for cell in ws_pos[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            ws_pos.auto_filter.ref = ws_pos.dimensions
            # color rows: Summary rows light grey, Operation rows white
            for i, row in enumerate(ws_pos.iter_rows(min_row=2, max_row=ws_pos.max_row), start=2):
                row_type_cell = ws_pos.cell(row=i, column=1)  # rowType in col A
                if str(row_type_cell.value).lower().startswith("summary"):
                    fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
                    for cell in row:
                        cell.fill = fill
                # operations left as default
        else:
            ws_pos["A1"] = "No POS operations found"

        # auto column width
        for sheet in [ws_acc, ws_stmt, ws_pos]:
            for col in sheet.columns:
                max_len = 0
                try:
                    col_letter = col[0].column_letter
                except Exception:
                    continue
                for cell in col:
                    try:
                        val = "" if cell.value is None else str(cell.value)
                        if len(val) > max_len:
                            max_len = len(val)
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)

        date_suffix = datetime.now().strftime("%Y-%m-%d_%H-%M")
        final_filename = f"{date_suffix}_{filename}"

        if self.excel_path:
            final_filename = str(self.excel_path.joinpath(final_filename))
            logging.log(msg="Final path: " + final_filename, level=logging.INFO)

        wb.save(final_filename)
        logging.log(msg=f"✅ Excel saved as: {final_filename}", level=logging.INFO)
        return final_filename

    def _setup_session(self):
        if self.config_jwt:
            self.session.headers["Authorization"] = f"Bearer {self.config_jwt}"

        if self.config_key:
            self.session.headers["apikey"] = self.config_key

        self.session.headers["User-agent"] = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                                              " (KHTML, like Gecko) Chrome/102.0.5005.49 Safari/537.36")
        self.session.headers["Accept"] = "application/json"
        self.session.headers["Content-Type"] = "application/json"

    def _gather_statements_rows(self, account_id: str, statements_obj: Dict[str, Any]) -> List[Dict[str, Any]]:
        ops = statements_obj.get("operations", []) or []
        rows = []

        #set max page count

        summary = {
            "accountNo": account_id,
            "openingBalance": statements_obj.get("openingBalance", 0),
            "closingBalance": statements_obj.get("closingBalance", 0),
            "availableOpeningBalance": statements_obj.get("availableOpeningBalance", 0),
            "availableClosingBalance": statements_obj.get("availableClosingBalance", 0),
            "message": _normalize_value(statements_obj.get("message", "")),
            "page_current": statements_obj.get("pagination", {}).get("currentPage"),
            "page_total": statements_obj.get("pagination", {}).get("totalPages"),
        }

        if not ops:
            r = {**summary}
            op_fields = [
                "operationDate", "transactionDate", "transactionNo", "transactionType",
                "transactionDescription", "transactionCurrency", "amountInTransactionCurrency",
                "amountInAccountCurrency", "amountInTransactionCurrencyAzn", "transactionFXRate",
                "openingBalance_op", "closingBalance_op", "openingAvlBalance", "closingAvlBalance",
                "afterOperationBalance", "afterOperationAvlBalance",
                "counterPartyName", "counterPartyId", "counterPartyTin", "counterPartyPin",
                "cardNo", "sourceSystem"
            ]
            for f in op_fields:
                r[f] = "N/A"
            # normalize numeric page values
            r["page_current"] = r.get("page_current") if r.get("page_current") is not None else "N/A"
            r["page_total"] = r.get("page_total") if r.get("page_total") is not None else "N/A"
            rows.append(r)
            return rows

        for op in ops:
            r = {
                "accountNo": account_id,
                "openingBalance": statements_obj.get("openingBalance", 0),
                "closingBalance": statements_obj.get("closingBalance", 0),
                "availableOpeningBalance": statements_obj.get("availableOpeningBalance", 0),
                "availableClosingBalance": statements_obj.get("availableClosingBalance", 0),
                "message": _normalize_value(statements_obj.get("message", "")),
                "page_current": statements_obj.get("pagination", {}).get("currentPage"),
                "page_total": statements_obj.get("pagination", {}).get("totalPages"),
                # operation fields
                "operationDate": _normalize_value(op.get("operationDate")),
                "transactionDate": _normalize_value(op.get("transactionDate")),
                "transactionNo": _normalize_value(op.get("transactionNo")),
                "transactionType": _normalize_value(op.get("transactionType")),
                "transactionDescription": _normalize_value(op.get("transactionDescription")),
                "transactionCurrency": _normalize_value(op.get("transactionCurrency")),
                "amountInTransactionCurrency": op.get("amountInTransactionCurrency") if op.get(
                    "amountInTransactionCurrency") is not None else "N/A",
                "amountInAccountCurrency": op.get("amountInAccountCurrency") if op.get(
                    "amountInAccountCurrency") is not None else "N/A",
                "amountInTransactionCurrencyAzn": op.get("amountInTransactionCurrencyAzn") if op.get(
                    "amountInTransactionCurrencyAzn") is not None else "N/A",
                "transactionFXRate": op.get("transactionFXRate") if op.get("transactionFXRate") is not None else "N/A",
                "openingBalance_op": op.get("openingBalance") if op.get("openingBalance") is not None else "N/A",
                "closingBalance_op": op.get("closingBalance") if op.get("closingBalance") is not None else "N/A",
                "openingAvlBalance": op.get("openingAvlBalance") if op.get("openingAvlBalance") is not None else "N/A",
                "closingAvlBalance": op.get("closingAvlBalance") if op.get("closingAvlBalance") is not None else "N/A",
                "afterOperationBalance": op.get("afterOperationBalance") if op.get(
                    "afterOperationBalance") is not None else "N/A",
                "afterOperationAvlBalance": op.get("afterOperationAvlBalance") if op.get(
                    "afterOperationAvlBalance") is not None else "N/A",
                "counterPartyName": _normalize_value(op.get("counterPartyName")),
                "counterPartyId": _normalize_value(op.get("counterPartyId")),
                "counterPartyTin": _normalize_value(op.get("counterPartyTin")),
                "counterPartyPin": _normalize_value(op.get("counterPartyPin")),
                "cardNo": _normalize_value(op.get("cardNo")),
                "sourceSystem": _normalize_value(op.get("sourceSystem"))
            }
            # normalize page numbers
            r["page_current"] = r.get("page_current") if r.get("page_current") is not None else "N/A"
            r["page_total"] = r.get("page_total") if r.get("page_total") is not None else "N/A"
            rows.append(r)
        return rows

    def _gather_pos_rows(self, account_id: str, pos_blocks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Build hybrid (B1) rows for POS sheet.
        For each posStatement block produce a 'Summary' row, then its operations as 'Operation' rows.
        """
        rows: List[Dict[str, Any]] = []
        for block in pos_blocks:
            # summary info (opening/closing balances are objects)
            opening = block.get("openingBalance", {}) or {}
            closing = block.get("closingBalance", {}) or {}
            terminal = block.get("terminalInfo", {}) or {}

            summary_row = {
                "rowType": "Summary",
                "accountNo": account_id,
                "terminalId": _normalize_value(terminal.get("id")),
                "terminalAddress": _normalize_value(terminal.get("address")),
                "opening_amountToReceive": opening.get("amountToReceive", "N/A"),
                "opening_transactionAmount": opening.get("transactionAmount", "N/A"),
                "opening_transactionCurrency": _normalize_value(opening.get("transactionCurrency")),
                "opening_cashBack": opening.get("cashBack", "N/A"),
                "opening_transactionFee": opening.get("transactionFee", "N/A"),
                "closing_amountToReceive": closing.get("amountToReceive", "N/A"),
                "closing_transactionAmount": closing.get("transactionAmount", "N/A"),
                "closing_transactionCurrency": _normalize_value(closing.get("transactionCurrency")),
                "closing_cashBack": closing.get("cashBack", "N/A"),
                "closing_transactionFee": closing.get("transactionFee", "N/A"),
            }
            # normalize string empties / None
            summary_row = {k: _normalize_value(v) for k, v in summary_row.items()}
            rows.append(summary_row)

            # operations list
            ops = block.get("posOperationEntityList", []) or []
            for op in ops:
                balance = op.get("balance", {}) or {}
                op_row = {
                    "rowType": "Operation",
                    "accountNo": account_id,
                    "terminalId": _normalize_value(terminal.get("id")),
                    "postingDate": _normalize_value(op.get("postingDate")),
                    "transactionDate": _normalize_value(op.get("transactionDate")),
                    "transactionTime": _normalize_value(op.get("transactionTime")),
                    "cardName": _normalize_value(op.get("cardName")),
                    "cardNumber": _normalize_value(op.get("cardNumber")),
                    "cardType": _normalize_value(op.get("cardType")),
                    "approvalCode": _normalize_value(op.get("approvalCode")),
                    "description": _normalize_value(op.get("description")),
                    "processingType": _normalize_value(op.get("processingType")),
                    "referenceNumber": _normalize_value(op.get("referenceNumber")),
                    "taksitCount": _normalize_value(op.get("taksitCount")),
                    # balance fields
                    "balance_amountToReceive": balance.get("amountToReceive", "N/A"),
                    "balance_cashBack": balance.get("cashBack", "N/A"),
                    "balance_transactionAmount": balance.get("transactionAmount", "N/A"),
                    "balance_transactionCurrency": _normalize_value(balance.get("transactionCurrency")),
                    "balance_transactionFee": balance.get("transactionFee", "N/A"),
                }
                op_row = {k: _normalize_value(v) for k, v in op_row.items()}
                rows.append(op_row)
        return rows

    def _make_request(self, url: str, method: str = "GET", params: Dict = None, retries: int = 3) -> Dict:

        for retry in range(1, retries + 1):
            try:
                if method.upper() == "POST":
                    resp = self.session.post(url, json=params, timeout=30)
                else:
                    resp = self.session.get(url, params=params, timeout=30)
                resp.raise_for_status()
                try:
                    data = resp.json()

                    return data
                except ValueError:
                    logging.log(msg="⚠️ Response is not JSON", level=logging.INFO)
                    return {}

            except (requests.Timeout, requests.ConnectionError) as e:

                if retry < retries:
                    logging.warning(f"⏳ Timeout/Connection error (attempt {retry}/{retries}): {e}")
                    time.sleep(2)
                    continue
                else:
                    logging.error(f"❌ Failed after {retries} attempts: {url}")
                    return {}

            except requests.RequestException as e:
                logging.log(msg=f"❌ Ошибка запроса: {e} -> {url}", level=logging.INFO)
                return {}

        return {}

    # ---------- Accounts ----------
    def _load_accounts(self) -> List[Dict[str, Any]]:
        base_url = self.base_url
        accounts_path = self.accounts_list_path
        url = f"{base_url}{accounts_path}"
        response = self._make_request(url, "GET", {"accountType": "CURRENT"})
        if isinstance(response, dict):
            if "accounts" in response and isinstance(response["accounts"], list):

                return response["accounts"]
            for v in response.values():
                if isinstance(v, list):
                    return v
            return []
        elif isinstance(response, list):
            return response
        else:
            return []

    # ---------- Statements ----------
    def get_current_statements(self, account_id: str, date_from:str, date_to:str) -> Dict[str, Any]:
        base_url = self.base_url
        path = self.stmt_path
        path = path.replace("{accountId}", account_id)
        url = f"{base_url}{path}"

        if self.page_max_count == 0:
            params = {
                "pageNumber": 1,
                "fromDate": date_from,
                "toDate": date_to
            }

            resp = self._make_request(url, "POST", params) or {}

            logging.log(msg="Resetting max page count ...", level=logging.INFO)
            self.page_max_count = resp.get("paginationMetaData", {}).get("totalPages") or 0

            logging.log(msg=f"First statement request: {resp}", level=logging.INFO)
            return {"data": None}
        else:
            self.current_page += 1

            params = {
                "pageNumber": self.current_page,
                "fromDate": date_from,
                "toDate": date_to
            }

            resp = self._make_request(url, "POST", params) or {}

            logging.log(msg=f"Current request: {resp}", level=logging.INFO)

            time.sleep(0.5)

            return {
                "operations": resp.get("operations", []),
                "openingBalance": resp.get("openingBalance", 0),
                "closingBalance": resp.get("closingBalance", 0),
                "availableOpeningBalance": resp.get("availableOpeningBalance", 0),
                "availableClosingBalance": resp.get("availableClosingBalance", 0),
                "pagination": resp.get("paginationMetaData", {}),
                "message": resp.get("message", "")
            }


    # ---------- POS operations with cursor-based pagination ----------
    def get_pos_operations(self, account_id: str) -> List[Dict[str, Any]]:
        pos_operations_path = self.pos_operations
        pos_operations_path = pos_operations_path.replace("{accountId}", account_id)

        url = f"{self.base_url}{pos_operations_path}"

        all_blocks: List[Dict[str, Any]] = []
        cursor: Optional[str] = None
        fetch_all = True  # per earlier decision P2: fetch all pages

        while True:
            params = {}
            if cursor:
                params["cursorToken"] = cursor
            # GET request with optional cursorToken
            resp = self._make_request(url, "GET", params) or {}
            data = resp.get("data", {}) or {}
            blocks = data.get("posStatementList", []) or []
            if blocks:
                all_blocks.extend(blocks)

            page_resp = resp.get("pageResponse", {}) or {}
            cursor = page_resp.get("cursorToken")
            # break if no cursor or fetch_all disabled
            if not cursor or not fetch_all:
                break

        return all_blocks

    # ---------- Utilities & normalization ----------

    def _gather_accounts_table(self, accounts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows = []
        for acc in accounts:
            row = {
                "accountNo": _normalize_value(acc.get("accountNo")),
                "iban": _normalize_value(acc.get("iban")),
                "customerNo": _normalize_value(acc.get("customerNo")),
                "currency": _normalize_value(acc.get("currency")),
                "availableBalance": acc.get("availableBalance", 0),
                "blockedAmount": acc.get("blockedAmount", 0),
                "currentBalance": acc.get("currentBalance", 0),
                "todayOpeningBalance": acc.get("todayOpeningBalance", 0),
                "todayIncome": acc.get("todayIncome", 0),
                "todayOutcome": acc.get("todayOutcome", 0),
                "accountOpenDate": _normalize_value(acc.get("accountOpenDate")),
                "accountStatus": _normalize_value(acc.get("accountStatus")),
                "branchCode": _normalize_value(acc.get("branchCode")),
                "branchName": _normalize_value(acc.get("branchName")),
                "bankCode": _normalize_value(acc.get("bankCode")),
                "accountCategory": _normalize_value(acc.get("accountCategory")),
                "hasPos": acc.get("hasPos", False),
                "hasCard": acc.get("hasCard", False),
                "hasCredit": acc.get("hasCredit", False),
                "tin": _normalize_value(acc.get("tin")),
                "creditIsAllowed": acc.get("creditIsAllowed", False),
                "debitIsAllowed": acc.get("debitIsAllowed", False),
                "accountType": _normalize_value(acc.get("accountType")),
            }
            rows.append(row)
        return rows

    def process_data(self, date_from:str, date_to:str, jwt: str, api_key: str):

        #Создать сессию перед запросами
        self.config_jwt = jwt
        self.config_key = api_key

        self._setup_session()

        logging.log(msg="Loading accounts ...", level=logging.INFO)
        accounts = self._load_accounts()
        if not accounts:
            logging.log(msg="No account found, stopping ...", level=logging.INFO)
            return False

        logging.info(msg=f"Current accounts: {accounts}")

        accounts_table = self._gather_accounts_table(accounts=accounts)

        # collect statements and pos rows
        all_statements_rows: List[Dict[str, Any]] = []
        all_pos_rows: List[Dict[str, Any]] = []

        for acc in accounts:

            acc_no = acc.get("accountNo")

            if self.page_max_count == 0:
                self.get_current_statements(acc_no, date_from, date_to)

            logging.log(msg="\n" + "=" * 40, level=logging.INFO)
            logging.log(msg=f"Processing account: {acc_no}", level=logging.INFO)
            logging.log(msg="=" * 40, level=logging.INFO)
            logging.log(msg=f"Date range: {date_from} - {date_to}", level=logging.INFO)
            logging.log(msg=f"Current page: {self.current_page} / {self.page_max_count}", level=logging.INFO)


            # Statements
            while self.current_page < self.page_max_count:
                statements_obj = self.get_current_statements(acc_no, date_from, date_to)

                if (statements_obj.get("message") is not None and
                        "there is no operations for the period" in statements_obj.get("message").lower()):

                    logging.log(msg="No statements found for this account.", level=logging.INFO)
                    continue

                stmt_rows = self._gather_statements_rows(account_id=acc_no, statements_obj=statements_obj)
                all_statements_rows.extend(stmt_rows)

            # POS blocks (with pagination)
            pos_blocks = self.get_pos_operations(acc_no)
            pos_rows = self._gather_pos_rows(acc_no, pos_blocks)
            all_pos_rows.extend(pos_rows)

            self.current_page = 0
            self.page_max_count = 0

        logging.log(msg="\nSaving report to Excel ...", level=logging.INFO)
        self.save_report(accounts_table, all_statements_rows, all_pos_rows, filename="pasha_report.xlsx")
        return True